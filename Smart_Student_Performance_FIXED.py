import tkinter as tk
from tkinter import ttk, messagebox
import os
import pickle
import pandas as pd
from openpyxl import Workbook, load_workbook
import requests

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

MODEL_FILE = os.path.join(
    BASE_DIR,
    "student_performance_model.pkl"
)

EXCEL_FILE = os.path.join(
    BASE_DIR,
    "student_prediction_data.xlsx"
)

# ============================================================
# N8N CLOUD PRODUCTION WEBHOOK
# ============================================================

N8N_WEBHOOK_URL = (
    "https://kerenn.app.n8n.cloud/webhook/"
    "student-performance-alert"
)

print("Using n8n Production Webhook:", N8N_WEBHOOK_URL)

# ============================================================
# LOAD ML MODEL
# ============================================================

if not os.path.exists(MODEL_FILE):
    messagebox.showerror(
        "Model Not Found",
        "student_performance_model.pkl was not found."
    )
    raise SystemExit

try:
    with open(MODEL_FILE, "rb") as file:
        model = pickle.load(file)
except Exception as error:
    messagebox.showerror(
        "Model Error",
        f"Could not load the ML model.\n\n{error}"
    )
    raise SystemExit


# ============================================================
# CALL N8N
# ============================================================

def call_n8n(
    student_id,
    student_name,
    student_email,
    attendance,
    study_hours,
    internal_marks,
    assignment_marks,
    previous_score,
    prediction,
    risk_level
):
    payload = {
        "student_id": str(student_id),
        "name": str(student_name),
        "email": str(student_email),
        "attendance": float(attendance),
        "study_hours": float(study_hours),
        "internal_marks": float(internal_marks),
        "assignment_completion": float(assignment_marks),
        "previous_performance": float(previous_score),
        "prediction": str(prediction),
        "risk": str(risk_level)
    }

    try:
        response = requests.post(
            N8N_WEBHOOK_URL,
            json=payload,
            timeout=120
        )
        response.raise_for_status()

    except requests.exceptions.Timeout:
        raise RuntimeError(
            "n8n request timed out.\n\n"
            "Make sure your n8n workflow is Published."
        )

    except requests.exceptions.ConnectionError:
        raise RuntimeError(
            "Could not connect to n8n Cloud.\n\n"
            "Check your internet connection and n8n Production URL."
        )

    except requests.exceptions.HTTPError as error:
        raise RuntimeError(
            f"n8n returned an HTTP error:\n\n{error}\n\n"
            f"URL:\n{N8N_WEBHOOK_URL}"
        )

    try:
        result = response.json()
    except ValueError:
        raise RuntimeError(
            "n8n did not return JSON.\n\n"
            f"Response received:\n{response.text[:500]}"
        )

    recommendation = ""

    if isinstance(result, dict):
        recommendation = (
            result.get("recommendation")
            or result.get("output")
            or result.get("text")
            or ""
        )

    elif isinstance(result, list) and len(result) > 0:
        first = result[0]
        if isinstance(first, dict):
            recommendation = (
                first.get("recommendation")
                or first.get("output")
                or first.get("text")
                or ""
            )

    recommendation = str(recommendation).strip()

    if not recommendation:
        raise RuntimeError(
            "n8n did not return a recommendation.\n\n"
            "Check the Respond to Webhook node.\n\n"
            f"n8n response:\n{result}"
        )

    return recommendation


# ============================================================
# SAVE TO EXCEL
# ============================================================

def save_to_excel(
    student_id,
    student_name,
    student_email,
    attendance,
    study_hours,
    internal_marks,
    assignment_marks,
    previous_score,
    prediction,
    risk_level,
    recommendation
):
    try:
        headers = [
            "Student ID",
            "Student Name",
            "Student Email",
            "Attendance (%)",
            "Study Hours",
            "Internal Marks (40)",
            "Assignment (%)",
            "Previous Score (%)",
            "Predicted Performance",
            "Risk Level",
            "Recommendation"
        ]

        if not os.path.exists(EXCEL_FILE):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Student Data"
            sheet.append(headers)
            workbook.save(EXCEL_FILE)

        workbook = load_workbook(EXCEL_FILE)

        if "Student Data" in workbook.sheetnames:
            sheet = workbook["Student Data"]
        else:
            sheet = workbook.create_sheet("Student Data")
            sheet.append(headers)

        # Add missing headers if an older Excel file is being used.
        current_headers = [
            sheet.cell(row=1, column=column).value
            for column in range(1, sheet.max_column + 1)
        ]

        for header in headers:
            if header not in current_headers:
                sheet.cell(
                    row=1,
                    column=sheet.max_column + 1
                ).value = header
                current_headers.append(header)

        row_data = {
            "Student ID": student_id,
            "Student Name": student_name,
            "Student Email": student_email,
            "Attendance (%)": attendance,
            "Study Hours": study_hours,
            "Internal Marks (40)": internal_marks,
            "Assignment (%)": assignment_marks,
            "Previous Score (%)": previous_score,
            "Predicted Performance": prediction,
            "Risk Level": risk_level,
            "Recommendation": recommendation
        }

        new_row = sheet.max_row + 1

        for column, header in enumerate(current_headers, start=1):
            if header in row_data:
                sheet.cell(
                    row=new_row,
                    column=column
                ).value = row_data[header]

        workbook.save(EXCEL_FILE)
        return True

    except PermissionError:
        messagebox.showerror(
            "Excel Error",
            "Please close student_prediction_data.xlsx "
            "and try again."
        )
        return False

    except Exception as error:
        messagebox.showerror(
            "Excel Error",
            f"Could not save data.\n\n{error}"
        )
        return False


# ============================================================
# VALIDATION
# ============================================================

def validate_student_id(event=None):
    value = student_id_entry.get().strip()

    if value == "":
        student_id_error.config(text="")
    elif not value.isdigit():
        student_id_error.config(text="Numbers only")
    else:
        student_id_error.config(text="")


def validate_student_name(event=None):
    value = student_name_entry.get().strip()

    if value == "":
        student_name_error.config(text="")
    elif not all(
        character.isalpha() or character.isspace()
        for character in value
    ):
        student_name_error.config(text="Letters only")
    else:
        student_name_error.config(text="")


def validate_student_email(event=None):
    value = student_email_entry.get().strip()

    if value == "":
        student_email_error.config(text="")
    elif "@" not in value or "." not in value:
        student_email_error.config(text="Invalid email")
    else:
        student_email_error.config(text="")


def validate_attendance(event=None):
    value = attendance_entry.get().strip()

    if value == "":
        attendance_error.config(text="")
        return

    try:
        number = float(value)

        if number < 0 or number > 100:
            attendance_error.config(text="Must be 0 - 100")
        else:
            attendance_error.config(text="")

    except ValueError:
        attendance_error.config(text="Numbers only")


def validate_study_hours(event=None):
    value = study_hours_entry.get().strip()

    if value == "":
        study_hours_error.config(text="")
        return

    try:
        number = float(value)

        if number < 0 or number > 24:
            study_hours_error.config(text="Must be 0 - 24")
        else:
            study_hours_error.config(text="")

    except ValueError:
        study_hours_error.config(text="Numbers only")


def validate_internal_marks(event=None):
    value = internal_marks_entry.get().strip()

    if value == "":
        internal_error.config(text="")
        return

    try:
        number = float(value)

        if number < 0 or number > 40:
            internal_error.config(text="Must be 0 - 40")
        else:
            internal_error.config(text="")

    except ValueError:
        internal_error.config(text="Numbers only")


def validate_assignment(event=None):
    value = assignment_entry.get().strip()

    if value == "":
        assignment_error.config(text="")
        return

    try:
        number = float(value)

        if number < 0 or number > 100:
            assignment_error.config(text="Must be 0 - 100")
        else:
            assignment_error.config(text="")

    except ValueError:
        assignment_error.config(text="Numbers only")


def validate_previous_score(event=None):
    value = previous_score_entry.get().strip()

    if value == "":
        previous_score_error.config(text="")
        return

    try:
        number = float(value)

        if number < 0 or number > 100:
            previous_score_error.config(text="Must be 0 - 100")
        else:
            previous_score_error.config(text="")

    except ValueError:
        previous_score_error.config(text="Numbers only")


# ============================================================
# PREDICTION
# ============================================================

def predict_performance():
    try:
        student_id = student_id_entry.get().strip()
        student_name = student_name_entry.get().strip()
        student_email = student_email_entry.get().strip()

        # --------------------------------------------------------
        # Required fields
        # --------------------------------------------------------

        if student_id == "":
            messagebox.showwarning(
                "Missing Information",
                "Please enter Student ID."
            )
            student_id_entry.focus()
            return

        if student_name == "":
            messagebox.showwarning(
                "Missing Information",
                "Please enter Student Name."
            )
            student_name_entry.focus()
            return

        if student_email == "":
            messagebox.showwarning(
                "Missing Information",
                "Please enter Student Email."
            )
            student_email_entry.focus()
            return

        # --------------------------------------------------------
        # Text validation
        # --------------------------------------------------------

        if not student_id.isdigit():
            messagebox.showerror(
                "Invalid Student ID",
                "Student ID must contain numbers only."
            )
            student_id_entry.focus()
            return

        if not all(
            character.isalpha() or character.isspace()
            for character in student_name
        ):
            messagebox.showerror(
                "Invalid Student Name",
                "Student Name must contain letters only."
            )
            student_name_entry.focus()
            return

        if "@" not in student_email or "." not in student_email:
            messagebox.showerror(
                "Invalid Email",
                "Please enter a valid Email ID."
            )
            student_email_entry.focus()
            return

        # --------------------------------------------------------
        # Numeric values
        # --------------------------------------------------------

        attendance = float(attendance_entry.get())
        study_hours = float(study_hours_entry.get())
        internal_marks = float(internal_marks_entry.get())
        assignment_marks = float(assignment_entry.get())
        previous_score = float(previous_score_entry.get())

        # --------------------------------------------------------
        # Range validation
        # --------------------------------------------------------

        if not 0 <= attendance <= 100:
            raise ValueError(
                "Attendance must be between 0 and 100."
            )

        if not 0 <= study_hours <= 24:
            raise ValueError(
                "Study Hours must be between 0 and 24."
            )

        if not 0 <= internal_marks <= 40:
            raise ValueError(
                "Internal Marks must be between 0 and 40."
            )

        if not 0 <= assignment_marks <= 100:
            raise ValueError(
                "Assignment must be between 0 and 100."
            )

        if not 0 <= previous_score <= 100:
            raise ValueError(
                "Previous Score must be between 0 and 100."
            )

        # --------------------------------------------------------
        # Create ML input
        # These names are kept from your original code.
        # --------------------------------------------------------

        new_student = pd.DataFrame(
            [[
                attendance,
                study_hours,
                internal_marks,
                assignment_marks,
                previous_score
            ]],
            columns=[
                "Attendance_Percentage",
                "Study_Hours",
                "IA_Marks",
                "Assignment_score",
                "Previous_score"
            ]
        )

        # --------------------------------------------------------
        # ML prediction
        # --------------------------------------------------------

        prediction = model.predict(new_student)
        performance = str(prediction[0])

        # --------------------------------------------------------
        # Risk level + local recommendation fallback
        # --------------------------------------------------------

        if performance.upper() in ["HIGH", "EXCELLENT"]:
            risk_level = "Low Risk"

            local_recommendation = (
                "Excellent performance. Continue the same study "
                "routine and participate actively in class."
            )

            risk_color = "#1976D2"

        elif performance.upper() in ["MEDIUM", "GOOD", "AVERAGE"]:
            risk_level = "Medium Risk"

            local_recommendation = (
                "Performance is satisfactory. Increase study hours "
                "and revise difficult topics regularly."
            )

            risk_color = "#D89D44"

        else:
            risk_level = "High Risk"

            local_recommendation = (
                "Needs improvement. Attend classes regularly, "
                "complete assignments, and seek help from teachers."
            )

            risk_color = "#D32F2F"

        # --------------------------------------------------------
        # Display ML result immediately
        # --------------------------------------------------------

        prediction_result.config(
            text=(
                f"Student: {student_name}\n"
                f"Predicted Performance: {performance}"
            ),
            foreground="#CC3982"
        )

        risk_result.config(
            text=f"Risk Level: {risk_level}",
            foreground=risk_color
        )

        recommendation_result.config(
            text="AI Recommendation: Generating through n8n..."
        )

        root.update_idletasks()

        # --------------------------------------------------------
        # Send data to n8n
        # Gemini + Gmail should happen inside n8n.
        # --------------------------------------------------------

        recommendation = call_n8n(
            student_id=student_id,
            student_name=student_name,
            student_email=student_email,
            attendance=attendance,
            study_hours=study_hours,
            internal_marks=internal_marks,
            assignment_marks=assignment_marks,
            previous_score=previous_score,
            prediction=performance,
            risk_level=risk_level
        )

        # --------------------------------------------------------
        # Display AI recommendation
        # --------------------------------------------------------

        recommendation_result.config(
            text=f"AI Recommendation: {recommendation}"
        )

        # --------------------------------------------------------
        # Save complete data to Excel
        # --------------------------------------------------------

        saved = save_to_excel(
            student_id,
            student_name,
            student_email,
            attendance,
            study_hours,
            internal_marks,
            assignment_marks,
            previous_score,
            performance,
            risk_level,
            recommendation
        )

        if saved:
            messagebox.showinfo(
                "Prediction Complete",
                f"Prediction completed successfully!\n\n"
                f"Student: {student_name}\n"
                f"Performance: {performance}\n"
                f"Risk: {risk_level}\n\n"
                f"✓ AI recommendation generated\n"
                f"✓ n8n workflow completed\n"
                f"✓ Data saved to Excel\n\n"
                f"Saved to:\n{EXCEL_FILE}"
            )

    except ValueError as error:
        messagebox.showerror(
            "Invalid Input",
            str(error)
        )

    except requests.exceptions.RequestException as error:
        messagebox.showerror(
            "n8n Error",
            f"Could not connect to n8n Cloud.\n\n{error}"
        )

    except Exception as error:
        messagebox.showerror(
            "Prediction Error",
            f"Something went wrong:\n\n{error}"
        )


# ============================================================
# CLEAR
# ============================================================

def clear_fields():
    entries = [
        student_id_entry,
        student_name_entry,
        student_email_entry,
        attendance_entry,
        study_hours_entry,
        internal_marks_entry,
        assignment_entry,
        previous_score_entry
    ]

    for entry in entries:
        entry.delete(0, tk.END)

    error_labels = [
        student_id_error,
        student_name_error,
        student_email_error,
        attendance_error,
        study_hours_error,
        internal_error,
        assignment_error,
        previous_score_error
    ]

    for label in error_labels:
        label.config(text="")

    prediction_result.config(text="")
    risk_result.config(text="")
    recommendation_result.config(text="")


# ============================================================
# EXIT
# ============================================================

def exit_application():
    if messagebox.askyesno(
        "Exit",
        "Do you want to exit the application?"
    ):
        root.destroy()


# ============================================================
# TKINTER WINDOW
# ============================================================

root = tk.Tk()

root.title(
    "Smart Student Performance Prediction System"
)

root.geometry("1050x850")
root.resizable(False, False)

root.configure(bg="#F2F6FB")

# ============================================================
# STYLES
# ============================================================

style = ttk.Style()
style.theme_use("clam")

style.configure(
    "Title.TLabel",
    background="#F2F6FB",
    foreground="#1F4E79",
    font=("Arial", 22, "bold")
)

style.configure(
    "Predict.TButton",
    font=("Arial", 11, "bold"),
    padding=8,
    background="#88DF8B",
    foreground="white"
)

style.configure(
    "Clear.TButton",
    font=("Arial", 11, "bold"),
    padding=8,
    background="#E2CD70",
    foreground="white"
)

style.configure(
    "Exit.TButton",
    font=("Arial", 11, "bold"),
    padding=8,
    background="#AD716C",
    foreground="white"
)

# ============================================================
# TITLE
# ============================================================

title_label = ttk.Label(
    root,
    text=(
        "SMART STUDENT PERFORMANCE "
        "PREDICTION SYSTEM"
    ),
    style="Title.TLabel",
    anchor="center"
)

title_label.pack(pady=15)

# ============================================================
# MAIN FRAME
# ============================================================

main_frame = tk.Frame(
    root,
    bg="#F2F6FB"
)

main_frame.pack(
    padx=30,
    fill="x"
)

main_frame.grid_columnconfigure(0, weight=1)
main_frame.grid_columnconfigure(1, weight=1)

# ============================================================
# STUDENT INFORMATION
# ============================================================

student_frame = tk.LabelFrame(
    main_frame,
    text="Student Information",
    bg="white",
    fg="#C67CE4",
    font=("Arial", 12, "bold"),
    padx=15,
    pady=12,
    width=480,
    height=230
)

student_frame.grid(
    row=0,
    column=0,
    padx=(0, 10),
    pady=6,
    sticky="nsew"
)

student_frame.grid_propagate(False)

# Student ID

tk.Label(
    student_frame,
    text="Student ID",
    bg="white",
    font=("Arial", 11)
).grid(
    row=0,
    column=0,
    padx=8,
    pady=8,
    sticky="w"
)

student_id_entry = ttk.Entry(
    student_frame,
    width=28
)

student_id_entry.grid(
    row=0,
    column=1,
    padx=8,
    pady=8
)

student_id_error = tk.Label(
    student_frame,
    text="",
    bg="white",
    fg="#D32F2F",
    font=("Arial", 8)
)

student_id_error.grid(
    row=0,
    column=2,
    padx=5,
    sticky="w"
)

# Student Name

tk.Label(
    student_frame,
    text="Student Name",
    bg="white",
    font=("Arial", 11)
).grid(
    row=1,
    column=0,
    padx=8,
    pady=8,
    sticky="w"
)

student_name_entry = ttk.Entry(
    student_frame,
    width=28
)

student_name_entry.grid(
    row=1,
    column=1,
    padx=8,
    pady=8
)

student_name_error = tk.Label(
    student_frame,
    text="",
    bg="white",
    fg="#D32F2F",
    font=("Arial", 8)
)

student_name_error.grid(
    row=1,
    column=2,
    padx=5,
    sticky="w"
)

# Email

tk.Label(
    student_frame,
    text="Student Email",
    bg="white",
    font=("Arial", 11)
).grid(
    row=2,
    column=0,
    padx=8,
    pady=8,
    sticky="w"
)

student_email_entry = ttk.Entry(
    student_frame,
    width=28
)

student_email_entry.grid(
    row=2,
    column=1,
    padx=8,
    pady=8
)

student_email_error = tk.Label(
    student_frame,
    text="",
    bg="white",
    fg="#D32F2F",
    font=("Arial", 8)
)

student_email_error.grid(
    row=2,
    column=2,
    padx=5,
    sticky="w"
)

# ============================================================
# ACADEMIC INFORMATION
# ============================================================

academic_frame = tk.LabelFrame(
    main_frame,
    text="Academic Information",
    bg="white",
    fg="#C67CE4",
    font=("Arial", 12, "bold"),
    padx=15,
    pady=12,
    width=480,
    height=280
)

academic_frame.grid(
    row=0,
    column=1,
    padx=(10, 0),
    pady=6,
    sticky="nsew"
)

academic_frame.grid_propagate(False)

# Attendance

tk.Label(
    academic_frame,
    text="Attendance (%)",
    bg="white",
    font=("Arial", 11)
).grid(
    row=0,
    column=0,
    padx=8,
    pady=6,
    sticky="w"
)

attendance_entry = ttk.Entry(
    academic_frame,
    width=25
)

attendance_entry.grid(
    row=0,
    column=1,
    padx=8,
    pady=6
)

attendance_error = tk.Label(
    academic_frame,
    text="",
    bg="white",
    fg="#D32F2F",
    font=("Arial", 8)
)

attendance_error.grid(
    row=0,
    column=2,
    padx=2,
    sticky="w"
)

# Study Hours

tk.Label(
    academic_frame,
    text="Study Hours per Day",
    bg="white",
    font=("Arial", 11)
).grid(
    row=1,
    column=0,
    padx=8,
    pady=6,
    sticky="w"
)

study_hours_entry = ttk.Entry(
    academic_frame,
    width=25
)

study_hours_entry.grid(
    row=1,
    column=1,
    padx=8,
    pady=6
)

study_hours_error = tk.Label(
    academic_frame,
    text="",
    bg="white",
    fg="#D32F2F",
    font=("Arial", 8)
)

study_hours_error.grid(
    row=1,
    column=2,
    padx=2,
    sticky="w"
)

# Internal Marks

tk.Label(
    academic_frame,
    text="Internal Marks (40)",
    bg="white",
    font=("Arial", 11)
).grid(
    row=2,
    column=0,
    padx=8,
    pady=6,
    sticky="w"
)

internal_marks_entry = ttk.Entry(
    academic_frame,
    width=25
)

internal_marks_entry.grid(
    row=2,
    column=1,
    padx=8,
    pady=6
)

internal_error = tk.Label(
    academic_frame,
    text="",
    bg="white",
    fg="#D32F2F",
    font=("Arial", 8)
)

internal_error.grid(
    row=2,
    column=2,
    padx=2,
    sticky="w"
)

# Assignment

tk.Label(
    academic_frame,
    text="Assignment (%)",
    bg="white",
    font=("Arial", 11)
).grid(
    row=3,
    column=0,
    padx=8,
    pady=6,
    sticky="w"
)

assignment_entry = ttk.Entry(
    academic_frame,
    width=25
)

assignment_entry.grid(
    row=3,
    column=1,
    padx=8,
    pady=6
)

assignment_error = tk.Label(
    academic_frame,
    text="",
    bg="white",
    fg="#D32F2F",
    font=("Arial", 8)
)

assignment_error.grid(
    row=3,
    column=2,
    padx=2,
    sticky="w"
)

# Previous Score

tk.Label(
    academic_frame,
    text="Previous Score (%)",
    bg="white",
    font=("Arial", 11)
).grid(
    row=4,
    column=0,
    padx=8,
    pady=6,
    sticky="w"
)

previous_score_entry = ttk.Entry(
    academic_frame,
    width=25
)

previous_score_entry.grid(
    row=4,
    column=1,
    padx=8,
    pady=6
)

previous_score_error = tk.Label(
    academic_frame,
    text="",
    bg="white",
    fg="#D32F2F",
    font=("Arial", 8)
)

previous_score_error.grid(
    row=4,
    column=2,
    padx=2,
    sticky="w"
)

# ============================================================
# LIVE VALIDATION
# ============================================================

student_id_entry.bind(
    "<KeyRelease>",
    validate_student_id
)

student_name_entry.bind(
    "<KeyRelease>",
    validate_student_name
)

student_email_entry.bind(
    "<KeyRelease>",
    validate_student_email
)

attendance_entry.bind(
    "<KeyRelease>",
    validate_attendance
)

study_hours_entry.bind(
    "<KeyRelease>",
    validate_study_hours
)

internal_marks_entry.bind(
    "<KeyRelease>",
    validate_internal_marks
)

assignment_entry.bind(
    "<KeyRelease>",
    validate_assignment
)

previous_score_entry.bind(
    "<KeyRelease>",
    validate_previous_score
)

# ============================================================
# BUTTONS
# ============================================================

button_frame = tk.Frame(
    root,
    bg="#F2F6FB"
)

button_frame.pack(pady=12)

predict_button = ttk.Button(
    button_frame,
    text="Predict Performance",
    style="Predict.TButton",
    command=predict_performance
)

predict_button.grid(
    row=0,
    column=0,
    padx=10
)

clear_button = ttk.Button(
    button_frame,
    text="Clear",
    style="Clear.TButton",
    command=clear_fields
)

clear_button.grid(
    row=0,
    column=1,
    padx=10
)

exit_button = ttk.Button(
    button_frame,
    text="Exit",
    style="Exit.TButton",
    command=exit_application
)

exit_button.grid(
    row=0,
    column=2,
    padx=10
)

# ============================================================
# PREDICTION RESULTS
# ============================================================

result_frame = tk.LabelFrame(
    root,
    text="Prediction Results",
    bg="white",
    fg="#1F4E79",
    font=("Arial", 12, "bold"),
    padx=20,
    pady=12,
    height=230
)

result_frame.pack(
    padx=30,
    pady=5,
    fill="x"
)

result_frame.pack_propagate(False)

prediction_result = tk.Label(
    result_frame,
    text="",
    bg="white",
    font=("Arial", 15, "bold"),
    justify="center"
)

prediction_result.pack(pady=3)

risk_result = tk.Label(
    result_frame,
    text="",
    bg="white",
    font=("Arial", 14, "bold")
)

risk_result.pack(pady=3)

tk.Label(
    result_frame,
    text="Recommendation:",
    bg="white",
    fg="#1F4E79",
    font=("Arial", 12, "bold")
).pack(
    pady=(5, 2)
)

recommendation_result = tk.Label(
    result_frame,
    text="",
    bg="white",
    font=("Arial", 11),
    wraplength=850,
    justify="center"
)

recommendation_result.pack(
    pady=3,
    padx=20
)

# ============================================================
# START
# ============================================================

root.mainloop()
