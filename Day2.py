import tkinter as tk
from tkinter import ttk, messagebox
import os
import pickle
import pandas as pd
from openpyxl import Workbook, load_workbook
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

MODEL_FILE = os.path.join(
    BASE_DIR,
    "student_performance_model.pkl"
)

EXCEL_FILE = os.path.join(
    BASE_DIR,
    "student_prediction_data.xlsx"
)

if not os.path.exists(MODEL_FILE):

    messagebox.showerror(
        "Model Not Found",
        "student_performance_model.pkl was not found.\n\n"
        "Please run d2 (1)(1).py first."
    )

    raise SystemExit


try:

    with open(MODEL_FILE, "rb") as file:
        model = pickle.load(file)

except Exception as error:

    messagebox.showerror(
        "Model Error",
        f"Could not load the ML model.\n\n{error}"
    )

    raise SystemExit
def save_to_excel(
    student_id,
    student_name,
    attendance,
    study_hours,
    internal_marks,
    assignment_marks,
    previous_score,
    predicted_performance,
    risk_level,
    recommendation
):

    try:
        if not os.path.exists(EXCEL_FILE):

            workbook = Workbook()

            sheet = workbook.active

            sheet.title = "Student Data"

            sheet.append([
                "Student ID",
                "Student Name",
                "Attendance (%)",
                "Study Hours",
                "Internal Marks (30)",
                "Assignment (%)",
                "Previous Score (%)",
                "Predicted Performance",
                "Risk Level",
                "Recommendation"
            ])

            workbook.save(EXCEL_FILE)
        workbook = load_workbook(EXCEL_FILE)

        if "Student Data" in workbook.sheetnames:

            sheet = workbook["Student Data"]

        else:

            sheet = workbook.create_sheet(
                "Student Data"
            )

            sheet.append([
                "Student ID",
                "Student Name",
                "Attendance (%)",
                "Study Hours",
                "Internal Marks (30)",
                "Assignment (%)",
                "Previous Score (%)",
                "Predicted Performance",
                "Risk Level",
                "Recommendation"
            ])
        sheet.append([
            student_id,
            student_name,
            attendance,
            study_hours,
            internal_marks,
            assignment_marks,
            previous_score,
            predicted_performance,
            risk_level,
            recommendation
        ])
        workbook.save(EXCEL_FILE)
        return True
    except PermissionError:

        messagebox.showerror(
            "Excel Error",
            "Please close student_prediction_data.xlsx "
            "and try again."
        )

        return False


    except Exception as error:

        messagebox.showerror(
            "Excel Error",
            f"Could not save data.\n\n{error}"
        )

        return False
def validate_student_id(event=None):

    value = student_id_entry.get().strip()

    if value == "":

        student_id_error.config(text="")

    elif not value.isdigit():

        student_id_error.config(
            text="Numbers only"
        )

    else:

        student_id_error.config(text="")


def validate_student_name(event=None):

    value = student_name_entry.get().strip()

    if value == "":

        student_name_error.config(text="")

    elif not all(
        character.isalpha() or character.isspace()
        for character in value
    ):

        student_name_error.config(
            text="Letters only"
        )

    else:

        student_name_error.config(text="")


def validate_attendance(event=None):

    value = attendance_entry.get().strip()

    if value == "":

        attendance_error.config(text="")
        return

    try:

        number = float(value)

        if number < 0 or number > 100:

            attendance_error.config(
                text="Must be 0 - 100"
            )

        else:

            attendance_error.config(text="")

    except ValueError:

        attendance_error.config(
            text="Numbers only"
        )


def validate_study_hours(event=None):

    value = study_hours_entry.get().strip()

    if value == "":

        study_hours_error.config(text="")
        return

    try:

        number = float(value)

        if number < 0 or number > 24:

            study_hours_error.config(
                text="Must be 0 - 24"
            )

        else:

            study_hours_error.config(text="")

    except ValueError:

        study_hours_error.config(
            text="Numbers only"
        )


def validate_internal_marks(event=None):

    value = internal_marks_entry.get().strip()

    if value == "":

        internal_error.config(text="")
        return

    try:

        number = float(value)

        if number < 0 or number > 30:

            internal_error.config(
                text="Must be 0 - 30"
            )

        else:

            internal_error.config(text="")

    except ValueError:

        internal_error.config(
            text="Numbers only"
        )


def validate_assignment(event=None):

    value = assignment_entry.get().strip()

    if value == "":

        assignment_error.config(text="")
        return

    try:

        number = float(value)

        if number < 0 or number > 100:

            assignment_error.config(
                text="Must be 0 - 100"
            )

        else:

            assignment_error.config(text="")

    except ValueError:

        assignment_error.config(
            text="Numbers only"
        )


def validate_previous_score(event=None):

    value = previous_score_entry.get().strip()

    if value == "":

        previous_score_error.config(text="")
        return

    try:

        number = float(value)

        if number < 0 or number > 100:

            previous_score_error.config(
                text="Must be 0 - 100"
            )

        else:

            previous_score_error.config(text="")

    except ValueError:

        previous_score_error.config(
            text="Numbers only"
        )
def predict_performance():

    try:
        student_id = student_id_entry.get().strip()

        student_name = student_name_entry.get().strip()


        if student_id == "" or student_name == "":

            messagebox.showwarning(
                "Missing Information",
                "Please enter Student ID and Student Name."
            )

            return
        if not student_id.isdigit():

            messagebox.showerror(
                "Invalid Student ID",
                "Student ID must contain numbers only."
            )

            student_id_entry.focus()

            return
        if not all(
            character.isalpha() or character.isspace()
            for character in student_name
        ):

            messagebox.showerror(
                "Invalid Student Name",
                "Student Name must contain letters only."
            )

            student_name_entry.focus()

            return
        attendance = float(
            attendance_entry.get()
        )

        study_hours = float(
            study_hours_entry.get()
        )

        internal_marks = float(
            internal_marks_entry.get()
        )

        assignment_marks = float(
            assignment_entry.get()
        )

        previous_score = float(
            previous_score_entry.get()
        )

        if not 0 <= attendance <= 100:

            raise ValueError(
                "Attendance must be between 0 and 100."
            )


        if not 0 <= study_hours <= 24:

            raise ValueError(
                "Study Hours must be between 0 and 24."
            )


        if not 0 <= internal_marks <= 30:

            raise ValueError(
                "Internal Marks must be between 0 and 30."
            )


        if not 0 <= assignment_marks <= 100:

            raise ValueError(
                "Assignment must be between 0 and 100."
            )


        if not 0 <= previous_score <= 100:

            raise ValueError(
                "Previous Score must be between 0 and 100."
            )

        new_student = pd.DataFrame(
            [[
                attendance,
                study_hours,
                internal_marks,
                assignment_marks,
                previous_score
            ]],
            columns=[
                "Attendance_Percentage",
                "Study_Hours",
                "IA_Marks",
                "Assignment_score",
                "Previous_score"
            ]
        )
        prediction = model.predict(
            new_student
        )
        performance = prediction[0]

        if performance == "High":

            risk_level = "Low Risk"

            recommendation = (
                "Excellent performance. "
                "Continue the same study routine "
                "and participate actively in class."
            )

            risk_color = "#1976D2"


        elif performance == "Medium":

            risk_level = "Medium Risk"

            recommendation = (
                "Performance is satisfactory. "
                "Increase study hours and revise "
                "difficult topics regularly."
            )

            risk_color = "#D89D44"


        else:

            risk_level = "High Risk"

            recommendation = (
                "Needs improvement. "
                "Attend classes regularly, "
                "complete assignments, and seek "
                "help from teachers."
            )

            risk_color = "#D32F2F"

        prediction_result.config(
            text=(
                f"Student: {student_name}\n"
                f"Predicted Performance: {performance}"
            ),
            foreground="#CC3982"
        )


        risk_result.config(
            text=f"Risk Level: {risk_level}",
            foreground=risk_color
        )


        recommendation_result.config(
            text=recommendation,
            foreground="#333333"
        )
        saved = save_to_excel(

            student_id,
            student_name,
            attendance,
            study_hours,
            internal_marks,
            assignment_marks,
            previous_score,
            performance,
            risk_level,
            recommendation

        )


        if saved:

            messagebox.showinfo(
                "Prediction Complete",

                f"Prediction completed!\n\n"
                f"Student: {student_name}\n"
                f"Performance: {performance}\n"
                f"Risk: {risk_level}\n\n"
                f"Saved to:\n"
                f"{EXCEL_FILE}"
            )


    except ValueError as error:

        messagebox.showerror(
            "Invalid Input",
            str(error)
        )


    except Exception as error:

        messagebox.showerror(
            "Prediction Error",
            f"Something went wrong:\n\n{error}"
        )

def clear_fields():

    entries = [

        student_id_entry,
        student_name_entry,
        attendance_entry,
        study_hours_entry,
        internal_marks_entry,
        assignment_entry,
        previous_score_entry

    ]

    for entry in entries:

        entry.delete(
            0,
            tk.END
        )


    prediction_result.config(
        text=""
    )

    risk_result.config(
        text=""
    )

    recommendation_result.config(
        text=""
    )

def exit_application():

    if messagebox.askyesno(
        "Exit",
        "Do you want to exit the application?"
    ):

        root.destroy()


root = tk.Tk()

root.title(
    "Smart Student Performance Prediction System"
)

root.geometry(
    "950x750"
)

root.resizable(
    False,
    False
)

root.configure(
    bg="#F2F6FB"
)

style = ttk.Style()

style.theme_use("clam")


style.configure(
    "Title.TLabel",
    background="#F2F6FB",
    foreground="#1F4E79",
    font=("Arial", 22, "bold")
)


style.configure(
    "Predict.TButton",
    font=("Arial", 11, "bold"),
    padding=8,
    background="#88DF8B",
    foreground="white"
)


style.configure(
    "Clear.TButton",
    font=("Arial", 11, "bold"),
    padding=8,
    background="#E2CD70",
    foreground="white"
)


style.configure(
    "Exit.TButton",
    font=("Arial", 11, "bold"),
    padding=8,
    background="#AD716C",
    foreground="white"
)

title_label = ttk.Label(

    root,

    text=(
        "SMART STUDENT PERFORMANCE "
        "PREDICTION SYSTEM"
    ),

    style="Title.TLabel",

    anchor="center"
)

title_label.pack(
    pady=20
)

main_frame = tk.Frame(
    root,
    bg="#F2F6FB"
)

main_frame.pack(
    padx=40,
    fill="x"
)


main_frame.grid_columnconfigure(
    0,
    weight=1
)

main_frame.grid_columnconfigure(
    1,
    weight=1
)
student_frame = tk.LabelFrame(

    main_frame,

    text="Student Information",

    bg="white",

    fg="#C67CE4",

    font=("Arial", 12, "bold"),

    padx=15,

    pady=15,

    width=400,

    height=200
)

student_frame.grid(

    row=0,

    column=0,

    padx=(0, 15),

    pady=6,

    sticky="nsew"
)

student_frame.grid_propagate(False)


tk.Label(

    student_frame,

    text="Student ID",

    bg="white",

    font=("Arial", 11)

).grid(

    row=0,

    column=0,

    padx=8,

    pady=10,

    sticky="w"
)


student_id_entry = ttk.Entry(

    student_frame,

    width=28

)

student_id_entry.grid(

    row=0,

    column=1,

    padx=8,

    pady=10
)


student_id_error = tk.Label(

    student_frame,

    text="",

    bg="white",

    fg="#D32F2F",

    font=("Arial", 8)

)

student_id_error.grid(

    row=0,

    column=2,

    padx=5,

    sticky="w"
)


tk.Label(

    student_frame,

    text="Student Name",

    bg="white",

    font=("Arial", 11)

).grid(

    row=1,

    column=0,

    padx=8,

    pady=10,

    sticky="w"
)


student_name_entry = ttk.Entry(

    student_frame,

    width=28

)

student_name_entry.grid(

    row=1,

    column=1,

    padx=8,

    pady=10
)


student_name_error = tk.Label(

    student_frame,

    text="",

    bg="white",

    fg="#D32F2F",

    font=("Arial", 8)

)

student_name_error.grid(

    row=1,

    column=2,

    padx=5,

    sticky="w"
)
academic_frame = tk.LabelFrame(

    main_frame,

    text="Academic Information",

    bg="white",

    fg="#C67CE4",

    font=("Arial", 12, "bold"),

    padx=20,

    pady=20,

    width=400,

    height=250
)

academic_frame.grid(

    row=0,

    column=1,

    padx=(15, 0),

    pady=5,

    sticky="nsew"
)

academic_frame.grid_propagate(False)


# Attendance

tk.Label(

    academic_frame,

    text="Attendance (%)",

    bg="white",

    font=("Arial", 11)

).grid(

    row=0,

    column=0,

    padx=8,

    pady=7,

    sticky="w"
)


attendance_entry = ttk.Entry(

    academic_frame,

    width=25

)

attendance_entry.grid(

    row=0,

    column=1,

    padx=8,

    pady=7
)


attendance_error = tk.Label(

    academic_frame,

    text="",

    bg="white",

    fg="#D32F2F",

    font=("Arial", 8)

)

attendance_error.grid(

    row=0,

    column=2,

    padx=2,

    sticky="w"
)


# Study Hours

tk.Label(

    academic_frame,

    text="Study Hours per Day",

    bg="white",

    font=("Arial", 11)

).grid(

    row=1,

    column=0,

    padx=8,

    pady=7,

    sticky="w"
)


study_hours_entry = ttk.Entry(

    academic_frame,

    width=25

)

study_hours_entry.grid(

    row=1,

    column=1,

    padx=8,

    pady=7
)


study_hours_error = tk.Label(

    academic_frame,

    text="",

    bg="white",

    fg="#D32F2F",

    font=("Arial", 8)

)

study_hours_error.grid(

    row=1,

    column=2,

    padx=2,

    sticky="w"
)


# Internal Marks

tk.Label(

    academic_frame,

    text="Internal Marks (30)",

    bg="white",

    font=("Arial", 11)

).grid(

    row=2,

    column=0,

    padx=8,

    pady=7,

    sticky="w"
)


internal_marks_entry = ttk.Entry(

    academic_frame,

    width=25

)

internal_marks_entry.grid(

    row=2,

    column=1,

    padx=8,

    pady=7
)


internal_error = tk.Label(

    academic_frame,

    text="",

    bg="white",

    fg="#D32F2F",

    font=("Arial", 8)

)

internal_error.grid(

    row=2,

    column=2,

    padx=2,

    sticky="w"
)


# Assignment

tk.Label(

    academic_frame,

    text="Assignment (%)",

    bg="white",

    font=("Arial", 11)

).grid(

    row=3,

    column=0,

    padx=8,

    pady=7,

    sticky="w"
)


assignment_entry = ttk.Entry(

    academic_frame,

    width=25

)

assignment_entry.grid(

    row=3,

    column=1,

    padx=8,

    pady=7
)


assignment_error = tk.Label(

    academic_frame,

    text="",

    bg="white",

    fg="#D32F2F",

    font=("Arial", 8)

)

assignment_error.grid(

    row=3,

    column=2,

    padx=2,

    sticky="w"
)


# Previous Score

tk.Label(

    academic_frame,

    text="Previous Score (%)",

    bg="white",

    font=("Arial", 11)

).grid(

    row=4,

    column=0,

    padx=8,

    pady=7,

    sticky="w"
)


previous_score_entry = ttk.Entry(

    academic_frame,

    width=25

)

previous_score_entry.grid(

    row=4,

    column=1,

    padx=8,

    pady=7
)


previous_score_error = tk.Label(

    academic_frame,

    text="",

    bg="white",

    fg="#D32F2F",

    font=("Arial", 8)

)

previous_score_error.grid(

    row=4,

    column=2,

    padx=2,

    sticky="w"
)
student_id_entry.bind(
    "<KeyRelease>",
    validate_student_id
)

student_name_entry.bind(
    "<KeyRelease>",
    validate_student_name
)

attendance_entry.bind(
    "<KeyRelease>",
    validate_attendance
)

study_hours_entry.bind(
    "<KeyRelease>",
    validate_study_hours
)

internal_marks_entry.bind(
    "<KeyRelease>",
    validate_internal_marks
)

assignment_entry.bind(
    "<KeyRelease>",
    validate_assignment
)

previous_score_entry.bind(
    "<KeyRelease>",
    validate_previous_score
)
button_frame = tk.Frame(

    root,

    bg="#F2F6FB"
)

button_frame.pack(
    pady=20
)


predict_button = ttk.Button(

    button_frame,

    text="Predict Performance",

    style="Predict.TButton",

    command=predict_performance
)

predict_button.grid(

    row=0,

    column=0,

    padx=10
)


clear_button = ttk.Button(

    button_frame,

    text="Clear",

    style="Clear.TButton",

    command=clear_fields
)

clear_button.grid(

    row=0,

    column=1,

    padx=10
)


exit_button = ttk.Button(

    button_frame,

    text="Exit",

    style="Exit.TButton",

    command=exit_application
)

exit_button.grid(
    row=0,
    column=2,
    padx=10
)
result_frame = tk.LabelFrame(
    root,
    text="Prediction Results",
    bg="white",
    fg="#1F4E79",
    font=("Arial", 12, "bold"),
    padx=20,
    pady=15,
   height=190
)
result_frame.pack(
    padx=35,
    pady=5,
    fill="x")
result_frame.pack_propagate(False)

prediction_result = tk.Label(
    result_frame,
    text="",
    bg="white",
    font=("Arial", 15, "bold"),
    justify="center"
)
prediction_result.pack(
    pady=5
)
risk_result = tk.Label(
    result_frame,
    text="",
    bg="white",
    font=("Arial", 14, "bold")
)
risk_result.pack(
    pady=5
)
tk.Label(
    result_frame,
    text="Recommendation:",
    bg="white",
    fg="#1F4E79",
    font=("Arial", 12, "bold")
).pack(
    pady=(5, 3)
)

recommendation_result = tk.Label(
    result_frame,
    text="",
    bg="white",
    font=("Arial", 11),
    wraplength=750,
    justify="center"
)
recommendation_result.pack(
    pady=3,
    padx=20
)
root.mainloop()
